from openpyxl import load_workbook, styles
from openpyxl.styles import Border, Side


def get_column_index(sheet, column: str):
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell_value in enumerate(row, 1):
            if cell_value == column:
                return idx
    return None


def highlight_mismatch_pair(excel_file, sheet):
    wb = load_workbook(excel_file)
    ws = wb[sheet]

    yellow_fill = styles.PatternFill(
        start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
    )

    left_border = Border(
        left=Side(border_style="thick"),
        right=Side(border_style="thin"),
        top=Side(border_style="thick"),
        bottom=Side(border_style="thick"),
    )
    right_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thick"),
        top=Side(border_style="thick"),
        bottom=Side(border_style="thick"),
    )

    for row in ws.iter_rows(
        min_row=2, min_col=0, max_row=ws.max_row, max_col=ws.max_column
    ):
        for col_idx in range(0, ws.max_column, 2):
            if row[col_idx].value != row[col_idx + 1].value:
                row[col_idx].fill = yellow_fill
                row[col_idx + 1].fill = yellow_fill

                row[col_idx].border = left_border
                row[col_idx + 1].border = right_border

    wb.save(excel_file)
